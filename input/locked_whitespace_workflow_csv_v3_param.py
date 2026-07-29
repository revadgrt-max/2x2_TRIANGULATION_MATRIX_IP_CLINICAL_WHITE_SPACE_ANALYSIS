#!/usr/bin/env python3
"""
locked_whitespace_workflow_csv_v3.py
=============================================================================
V3 of locked_whitespace_workflow_csv.py. The change vs v2 is WHO decides the x-axis
(FTO / epitope crowding) and what the LLM is allowed to do:

    v1: x = crowding_from_count(claims-grounded family count)       — a COUNT proxy
    v2: x = LLM EPITOPE JUDGE over each target's claim text         — the model IS the axis
    v3: x = DETERMINISTIC WEIGHTED THREE-LAYER FTO  +  one-way LLM FLOOR CHECK

v3 x-axis contract (the decider is deterministic; the LLM can only ratchet UP):
  X_det = max( 0.55*L1 + 0.25*L2 + 0.20*L3 , L1 )
    L1  EPITOPE / binding    target name, ALL modalities (cross-modality)   weight 0.55  + HARD FLOOR
    L2  FORMAT / conjugation target x THIS modality                          weight 0.25
    L3  USE / combination    target x modality, CONFINED TO INDICATION       weight 0.20
                             (+IO nodes -> combination claim space)
  Each layer is a log10(count)/scale density of its own claims-grounded patent slice, so
  "same CSV + same query + same weights/scales -> same X" holds, every run.

  THEN the LLM acts ONLY as a one-way FLOOR CHECK over the verbatim claim text:
    pull cohort -> assemble CLAIM text -> epitope.judge_epitope_crowding (LLM reasoning)
    -> quote-faithfulness guard -> epitope.verified_epitope_flag (DETERMINISTIC gate).
  A gated, evidence-backed 'locked' verdict raises X to a minimum:  X = max(X_det, floor).
  Open / abstain / unbacked / no-client / error  ->  X = X_det (the LLM can NEVER lower X).
  Because the deterministic metric always produces X, the LLM is OPTIONAL (REQUIRE_LLM=False):
  with no client the run completes and the floor check is simply skipped and flagged.

    y-axis = clinical performance   (verified, cited trial readouts)        low=failed -> high=validated
    x-axis = FTO / epitope crowding (deterministic weighted FTO + LLM floor) low=open   -> high=crowded
    Quadrants: TRUE WHITE SPACE | BATTLEGROUND | R&D TRAP | RED FLAGS   (top-2 = shortlist)

-----------------------------------------------------------------------------
THE SPLIT (this is the whole point of the file)

  PART 1  DETERMINISTIC CORE      same inputs -> same output, every run. Pure, no
                                  network, no model. These are GUARANTEES, and they
                                  are the ONLY things that decide the score/quadrant:
                                    date_filter   — anti-leakage (no post-cutoff data)
                                    score_target  — clinical-performance math (y)
                                    crowding_from_count — IP-density transform (x)
                                    quadrant / build_map1 — the 2x2 placement
                                  Plus the deterministic GATE: an outcome with no
                                  citation cannot be decisive (anti-fabrication).

  PART 2  NON-DETERMINISTIC       I/O + judgment. Injected as plug-ins so every
          PLUG-INS                external call is mockable and auditable. NOTHING
                                  here decides a score — each only produces an INPUT:
                                    pull_trials_fn        [I/O]   CT.gov census
                                    adjudicate_outcome_fn [JUDGE] read the publication
                                                                  -> (endpoint_met, citation)
                                    pull_patents_fn       [I/O]   PatSnap landscape
                                  A language instruction is a request; the Part-1
                                  gate is a guarantee — an LLM call cannot enter the
                                  score unless it carries a citation.

  PART 3  ORCHESTRATOR            run_node(): the fixed per-node order. Dependency
                                  injection; the deterministic functions gate the result.

  PART 4  OUTPUTS                 4-quadrant graph + 3 spreadsheets.
  PART 5  REGISTRY + main         the locked NSCLC 2L node set (TROP2 split 3 ways)
                                  and the wiring of the real adapters.

Run:  python3 locked_whitespace_workflow_csv.py
Data: this is the LOCAL-CSV variant of locked_whitespace_workflow.py. The two I/O
      plug-ins read local CSVs instead of hitting the network — everything else
      (the deterministic core, the JUDGE overlay, the orchestrator) is identical:
        trial census     <- extended_trial_master_with_outcomes_evidence.csv  (was CT.gov)
        patent landscape <- modality_agnostic_ip_master.csv                   (was PatSnap)
      Override the paths with env TRIAL_MASTER_CSV / IP_MASTER_CSV if needed.
"""
from __future__ import annotations

import csv
import datetime as dt
import math
import os
import re
import sys
from dataclasses import dataclass, field, asdict
from enum import Enum
from typing import Callable, Optional

TUMOR, LINE = "NSCLC", "2L"   # defaults; override at run time with --tumor / --line
AS_OF = "2026-05-05"          # leakage cutoff; override with --as-of
AS_OF_YEAR = 2026
TRIAL_CONDITION = None        # CT.gov condition scope; None -> derived from TUMOR. Override with --condition
_HERE = os.path.dirname(os.path.abspath(__file__))
OUTDIR = os.path.join(_HERE, "results_v3")   # default output dir; override with --outdir

# v2: the LLM epitope-reasoning layer (whitespace_pipeline.py / epitope.py contract).
sys.path.insert(0, _HERE)
from epitope import judge_epitope_crowding, verified_epitope_flag, EpitopeVerdict  # noqa: E402

JUDGE_MAX_PATENTS = 14       # cap claim records sent to the judge per target (cost/context)

# Local data sources that replace the two live API pulls (CT.gov / PatSnap).
TRIAL_MASTER_CSV = os.environ.get(
    "TRIAL_MASTER_CSV", os.path.join(_HERE, "extended_trial_master_with_outcomes_evidence.csv"))
def _first_existing(*paths):
    """Return the first path that exists, else the last (so the loader reports a clear miss)."""
    for p in paths:
        if p and os.path.exists(p):
            return p
    return paths[-1]


IP_MASTER_CSV = os.environ.get(
    "IP_MASTER_CSV",
    # The FINAL_NEWV2 claims-enriched master. The workspace was reorganised, so we try the
    # new "FINAL data sets/database_ip/" home first, then the old flat locations. The data
    # table is the same 2794-patent / 2793-claims-pulled set. _load_ip_master() is
    # preamble-tolerant, so a multi-section workbook dump also works if dropped in unmodified.
    _first_existing(
        os.path.join(_HERE, "FINAL data sets", "database_ip",
                     "FINAL_NEWV2_modality_agnostic_ip_master_claims_fulltext_enriched.DATA.csv"),
        os.path.join(_HERE, "FINAL_NEWV2_modality_agnostic_ip_master_claims_fulltext_enriched.DATA.csv"),
        os.path.join(_HERE, "ip_master_claims_fulltext_enriched_FINALV2.csv"),
    ))

csv.field_size_limit(10 ** 9)   # trial-master free-text fields are large


# ===========================================================================
# PART 1 — DETERMINISTIC CORE   (guarantees; pure; the ONLY deciders)
# ===========================================================================

# --- 1.0  anti-leakage guarantee -------------------------------------------
def date_filter(records: list[dict], *, as_of: str, date_key: str) -> list[dict]:
    """Keep only records dated on/before `as_of`; drop (and COUNT) post-cutoff and
    undated. A silent filter is one you can't trust — so it reports what it dropped.
    This is a guarantee, not a request: no post-cutoff data can enter the snapshot."""
    cutoff = _parse(as_of)
    kept, late, undated = [], 0, 0
    for r in records:
        d = _parse(r.get(date_key))
        if d is None:
            undated += 1; continue
        if d > cutoff:
            late += 1; continue
        kept.append(r)
    print(f"    [date_filter:{date_key}] as_of {as_of}: kept {len(kept)}, "
          f"dropped {late} post-cutoff, {undated} undated", file=sys.stderr)
    return kept


def _parse(s) -> Optional[dt.date]:
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y%m%d", "%Y-%m", "%Y"):
        try:
            return dt.datetime.strptime(str(s)[:10], fmt).date()
        except ValueError:
            continue
    return None


# --- 1.1  clinical-performance math (y) ------------------------------------
class Phase(Enum):
    PH1 = 1; PH2 = 2; PH3 = 3; APPROVAL = 4


class Outcome(Enum):
    POSITIVE = "positive"; NEGATIVE = "negative"; PENDING = "pending"; EXCLUDE = "exclude"


@dataclass
class Readout:
    """The unit of evidence. `endpoint_met` + `citation` are filled by the JUDGE
    plug-in (Part 2). The deterministic GATE below: no citation -> not decisive."""
    asset: str
    indication: str
    phase: Phase
    status: str
    start_year: Optional[int]
    combo: bool
    primary_endpoint: Optional[str]
    endpoint_met: Optional[bool]
    approved_or_commercial: bool
    why_stopped: Optional[str]
    citation: Optional[str]
    nct_id: Optional[str] = None
    has_results: bool = False
    superseded: bool = False


ENDPOINT_STRENGTH = {"OS": 1.0, "PFS": 0.7, "ORR": 0.5, "other": 0.3, None: 0.3}
PHASE_WEIGHT = {Phase.PH1: 0.15, Phase.PH2: 0.45, Phase.PH3: 0.85, Phase.APPROVAL: 1.0}
STALE_YEARS, PH3_FAIL_EMPHASIS, COVERAGE_K, SQUASH = 5, 1.30, 3, 1.5


def adjudicate(r: Readout, as_of_year: int) -> Outcome:
    """Status + cited publication -> Outcome. The GATE: a COMPLETED trial with an
    endpoint result is decisive ONLY if it carries a citation; otherwise PENDING."""
    if r.approved_or_commercial:
        return Outcome.POSITIVE
    if r.status == "WITHDRAWN":
        return Outcome.EXCLUDE
    if r.why_stopped in ("business", "accrual", "strategic", "funding", "enrollment"):
        return Outcome.EXCLUDE
    if r.why_stopped in ("efficacy", "safety", "futility", "toxicity"):
        return Outcome.NEGATIVE
    if r.status == "COMPLETED":
        if r.endpoint_met is True and r.citation:
            return Outcome.POSITIVE
        if r.endpoint_met is False and r.citation:
            return Outcome.NEGATIVE
        return Outcome.PENDING                 # no cited result -> abstain (the gate)
    if r.start_year is not None and (as_of_year - r.start_year) > STALE_YEARS:
        return Outcome.NEGATIVE                 # stale-incomplete = silent failure
    return Outcome.PENDING


def _basis_strength(r: Readout) -> float:
    if r.approved_or_commercial:
        return 1.0
    if r.endpoint_met is not None:
        return ENDPOINT_STRENGTH.get(r.primary_endpoint, 0.3)
    return 0.4


def contribution(r: Readout, o: Outcome) -> Optional[float]:
    if o in (Outcome.PENDING, Outcome.EXCLUDE):
        return None
    sign = 1.0 if o is Outcome.POSITIVE else -1.0
    phase = Phase.APPROVAL if r.approved_or_commercial else r.phase
    attr = 1.0 if not r.combo else (0.85 if sign > 0 else 0.45)   # combo loss can't indict the target alone
    w = PHASE_WEIGHT[phase] * _basis_strength(r) * attr
    if sign < 0:
        if r.phase is Phase.PH3:
            w *= PH3_FAIL_EMPHASIS
        if r.superseded:
            w *= 0.5
    return sign * w


@dataclass
class TargetResult:
    performance: Optional[float]; label: str; confidence: float
    decisive_n: int; detail: list


def _norm(s: str) -> str:
    return " ".join(sorted(s.lower().replace("-", " ").split()))


def score_target(readouts: list[Readout], *, as_of_year: int,
                 trial_volume: Optional[int] = None) -> TargetResult:
    """DETERMINISTIC: same Readouts -> same y. Abstains (None) without a decisive readout."""
    detail, bucket = [], {}
    decisive_n = 0; any_approval = neg_any = False
    max_phase = 0; oldest = None
    for r in readouts:
        o = adjudicate(r, as_of_year); c = contribution(r, o)
        detail.append({"asset": r.asset, "nct": r.nct_id, "phase": r.phase.name,
                       "outcome": o.value, "contribution": None if c is None else round(c, 3),
                       "citation": r.citation})
        max_phase = max(max_phase, r.phase.value)
        if r.start_year is not None:
            oldest = r.start_year if oldest is None else min(oldest, r.start_year)
        if c is not None:
            decisive_n += 1
            b = bucket.setdefault(_norm(r.indication), {"net": 0.0, "approved": False})
            b["net"] += c
            if r.approved_or_commercial:
                b["approved"] = True; any_approval = True
            neg_any |= c < 0
    if decisive_n == 0:
        return TargetResult(None, "ABSTAIN - no decisive readout", 0.0, 0, detail)
    ind = {k: (0.92 if b["approved"] else 0.5 + 0.5 * math.tanh(b["net"] / SQUASH))
           for k, b in bucket.items()}
    perf = 0.60 * max(ind.values()) + 0.40 * (sum(ind.values()) / len(ind))
    stalled = (not any_approval and max_phase < Phase.PH3.value and
               ((oldest is not None and oldest < 2018) or (trial_volume and trial_volume >= 50)))
    if any_approval:
        perf = max(perf, 0.75)
    if stalled:
        perf = min(perf, 0.33)
    conf = min(1.0, decisive_n / COVERAGE_K)
    if any_approval:    label = "VALIDATED (approved)"
    elif stalled:       label = "STALLED - no Ph3 in extensive history"
    elif perf >= 0.60:  label = "VALIDATED"
    elif not neg_any:   label = "PROMISING (thin)" if conf < 0.6 else "VALIDATED"
    elif perf <= 0.40:  label = "FAILED (late-phase)"
    else:               label = "CONTESTED"
    return TargetResult(round(perf, 3), label, round(conf, 2), decisive_n, detail)


# --- 1.2  patent-density transform (x) -------------------------------------
# `scale` is the log-divisor that sets what count = full crowding (1.0).
#   scale=5.0  -> broad family counts           (1->0 · ~100->0.4 · ~1k->0.6 · >=100k->1.0)
#   scale=3.0  -> CLAIMS-GROUNDED counts (FINALV2, ~full coverage; counts ~3-60)
#                 (1->0 · ~10->0.33 · ~30->0.49 · ~50->0.57 · ~1000->1.0)
# The claims-grounded cohort is ~1 order smaller than the broad one, so it needs its own
# scale or every target collapses to one side of the 0.5 split. Change it in ONE place.
CLAIMS_CROWDING_SCALE = 3.0


def crowding_from_count(total: Optional[int], *, scale: float = 5.0) -> Optional[float]:
    """Transparent, deterministic IP-density -> [0,1]: log10(total)/scale, clamped.
    None passes through."""
    if total is None:
        return None
    return round(max(0.0, min(1.0, math.log10(max(total, 1)) / scale)), 3)


# --- 1.2b  v3 DETERMINISTIC WEIGHTED THREE-LAYER FTO (the X-axis decider) ----
# v3 contract: X is set by a deterministic, reproducible weighted blend of three
# claim-grounded patent-density layers, each scoped DIFFERENTLY. The LLM (Part 2) is
# demoted to a one-way floor check on top — it can only ratchet X up, never set it.
#
#   L1  EPITOPE / binding   scope: target name, ALL modalities (cross-modality risk)
#                           signal: locking claims (epitope/domain/competition) — a broad
#                           "any antibody that binds <target>" / "competes with mAb X" claim
#                           blocks every modality, so this layer is shared across a target's
#                           nodes AND is enforced as a HARD FLOOR.                     weight 0.55
#   L2  FORMAT / conjugation scope: target x THIS node's modality                      weight 0.25
#   L3  USE / combination   scope: target x modality, CONFINED TO THE INDICATION
#                           (and, for +IO nodes, to combination claim space)            weight 0.20
#
#   X_det = max( 0.55*L1 + 0.25*L2 + 0.20*L3 ,  L1 )      # L1 also a hard floor
#
# Each layer is a log10(count)/scale density of its own claims-grounded patent slice, so
# "same CSV + same query + same weights -> same X" holds. Tune weights/scales in ONE place.
FTO_W_L1, FTO_W_L2, FTO_W_L3 = 0.55, 0.25, 0.20      # epitope / format / use weights (sum=1.0)
FTO_L1_IS_HARD_FLOOR = True                          # a locked epitope can't read as open
FTO_SCALE_L1, FTO_SCALE_L2, FTO_SCALE_L3 = 2.0, 1.8, 1.3   # per-layer log10-divisors (count->1.0)
FOCUS_MAX_TARGETS = 3   # BISPEC-FIX: bispecific-pair cohorts exclude multi-antigen 'list' patents (>3 distinct antigens)

# L3 helpers: which patents count as "use / combination", and how to spot the indication
# and an IO-combination context inside the enriched master's free-text columns.
_USE_CATEGORIES = {"treatment / regimen / combination"}
_INDICATION_TERMS = ("lung", "nsclc", "non-small", "non small")
_IO_COMBO_TERMS = ("durvalumab", "pembroli", "nivolu", "atezoli", "pd-1", "pd1",
                   "pd-l1", "pdl1", "checkpoint", "immunotherap", " io ")


def fto_layer_density(count: Optional[int], *, scale: float) -> float:
    """Deterministic per-layer density in [0,1]: log10(count)/scale, clamped. 0 -> 0.0."""
    if not count:
        return 0.0
    return round(max(0.0, min(1.0, math.log10(count) / scale)), 3)


# --- 1.3  node model + 2x2 placement ---------------------------------------
PERF_SPLIT = CROWDING_SPLIT = 0.5


def quadrant(validation: Optional[float], crowding: Optional[float]) -> str:
    if validation is None or crowding is None:
        return "UNRESOLVED"                          # missing axis -> never silently placed
    hi, crowded = validation >= PERF_SPLIT, crowding >= CROWDING_SPLIT
    if hi and not crowded:     return "TRUE WHITE SPACE"
    if hi and crowded:         return "BATTLEGROUND"
    if not hi and not crowded: return "R&D TRAP"
    return "RED FLAGS"


@dataclass
class Node:
    target: str; modality: str; configuration: str; population: str
    asset_examples: str = ""; biomarker: Optional[str] = None
    biomarker_prevalence: Optional[float] = None
    expression_prevalence_pct: Optional[float] = None
    # y
    validation_score: Optional[float] = None; validation_label: Optional[str] = None
    validation_confidence: float = 0.0
    trials_total: int = 0; trials_kept_as_of: int = 0
    trials_by_phase: dict = field(default_factory=dict)
    ph3_failure: Optional[bool] = None; discovery_complete: bool = False
    # x — IP/FTO density proxy. Honest name: target×modality FAMILY density,
    #     NOT epitope-resolved crowding.
    target_modality_crowding: Optional[float] = None; crowding_status: str = "NOT_PULLED"
    crowding_basis: Optional[str] = None         # v2: "llm_epitope_judge"
    epitope_crowding: Optional[float] = None     # v3: THE x-axis — FINAL X (weighted FTO, possibly floored up by LLM)
    epitope_confidence: Optional[float] = None   # judge confidence (gated) when a floor was applied
    epitope_rationale: Optional[str] = None      # judge rationale (audit)
    cited_claims: Optional[str] = None           # patents the verdict cited (audit)
    epitope_resolved: Optional[bool] = None      # True if a deterministic X was produced; else FLAGGED
    # v3 deterministic weighted FTO breakdown (audit) — the X decider
    fto_l1_epitope: Optional[float] = None       # L1 epitope/binding (target, all modalities)  [hard floor]
    fto_l2_format: Optional[float] = None        # L2 format/conjugation (target x modality)
    fto_l3_use: Optional[float] = None           # L3 use/combination (target x modality x indication)
    x_deterministic: Optional[float] = None      # X_det before any LLM floor check
    floor_applied: Optional[bool] = None         # True if the LLM floor check raised X above X_det
    floor_delta: Optional[float] = None          # X_final - X_det (>=0; LLM is one-way)
    fto_l1_lock_count: Optional[int] = None      # patents behind L1 (locking epitope claims)
    fto_l3_use_count: Optional[int] = None       # patents behind L3 (indication-confined use/combo)
    patent_family_count: Optional[int] = None    # claims-grounded count (context, not the x)
    patent_family_count_broad: Optional[int] = None  # transparency: all target×modality matches
    claims_coverage: Optional[float] = None      # grounded / broad — claim coverage of the cohort
    patent_recent: Optional[int] = None; patent_prior: Optional[int] = None
    patent_velocity_ratio: Optional[float] = None; patent_velocity_trend: Optional[str] = None
    most_recent_filing_year: Optional[int] = None
    key_patent: Optional[str] = None; blocking_assignee: Optional[str] = None
    cohort_patent_ids: list = field(default_factory=list)   # stable anchor set for claims-text pull
    cohort_size: Optional[int] = None
    ip_query: Optional[str] = None               # local FINAL_NEWV2 boolean match string (not a PatSnap call)
    # derived
    node_score: Optional[float] = None; quadrant: Optional[str] = None
    citations: dict = field(default_factory=dict)

    @property
    def label(self) -> str:
        return f"{self.target} {self.modality} ({self.configuration}/{self.population})"


def assemble_node_score(n: Node) -> Node:
    n.node_score = None if n.validation_score is None else round(n.validation_score, 3)
    n.quadrant = quadrant(n.validation_score, n.epitope_crowding)   # v2: x = LLM epitope read
    return n


def coverage_gaps(nodes: list[Node]) -> list[dict]:
    out = []
    for n in nodes:
        miss = []
        if not n.discovery_complete:                        miss.append("trial census not exhaustive")
        if n.crowding_status in ("NOT_PULLED", "UNREACHED"): miss.append("patent/FTO not reached")
        if str(n.crowding_status).startswith("UNRESOLVED"):
            miss.append(f"epitope read UNRESOLVED ({n.crowding_status}) — attempted, not resolved")
        if n.validation_score is None:                      miss.append("no decisive clinical readout (Y abstain)")
        if miss:
            out.append({"node": n.label, "gaps": "; ".join(miss)})
    return out


ALL_QUADRANTS = ("TRUE WHITE SPACE", "BATTLEGROUND", "R&D TRAP", "RED FLAGS")
SHORTLIST_QUADRANTS = ("TRUE WHITE SPACE", "BATTLEGROUND")


@dataclass
class Matrix2x2:
    cells: dict; unresolved: list

    def summary(self) -> dict:
        s = {q: len(self.cells.get(q, [])) for q in ALL_QUADRANTS}
        s["UNRESOLVED"] = len(self.unresolved); return s

    def shortlist(self) -> list:
        out = [n for q in SHORTLIST_QUADRANTS for n in self.cells.get(q, [])]
        return sorted(out, key=lambda n: -(n.node_score or 0.0))


def build_map1(nodes: list[Node]) -> Matrix2x2:
    cells = {q: [] for q in ALL_QUADRANTS}; unresolved = []
    for n in nodes:
        assemble_node_score(n)
        (cells[n.quadrant].append(n) if n.quadrant in cells else unresolved.append(n))
    for q in cells:
        cells[q].sort(key=lambda n: -(n.node_score or 0.0))
    return Matrix2x2(cells, unresolved)


# ===========================================================================
# PART 2 — NON-DETERMINISTIC PLUG-INS   (I/O + judgment; injected; produce INPUTS only)
# ===========================================================================

# --- 2a  [I/O]  trial census from the LOCAL trial-master CSV ----------------
# Was: ClinicalTrials.gov v2 API. Now: a search over the local trial-master CSV.
# Same dict shape out (incl. start_date, so the Part-1 date_filter is unchanged).
_PHASE_MAP = {"PHASE1": Phase.PH1, "EARLY_PHASE1": Phase.PH1, "PHASE2": Phase.PH2,
              "PHASE3": Phase.PH3, "PHASE4": Phase.APPROVAL}

_TRIAL_ROWS: Optional[list[dict]] = None


def _load_trial_master() -> list[dict]:
    """Read (and cache) the trial-master CSV once; the file is large so we keep the
    parsed rows in memory across all nodes. Missing file -> [] (flagged later)."""
    global _TRIAL_ROWS
    if _TRIAL_ROWS is None:
        try:
            with open(TRIAL_MASTER_CSV, newline="", encoding="utf-8") as f:
                _TRIAL_ROWS = list(csv.DictReader(f))
            print(f"    [trial_csv] loaded {len(_TRIAL_ROWS)} rows from "
                  f"{os.path.basename(TRIAL_MASTER_CSV)}", file=sys.stderr)
        except Exception as e:
            print(f"    [trial_csv] WARN cannot read {TRIAL_MASTER_CSV}: {e}", file=sys.stderr)
            _TRIAL_ROWS = []
    return _TRIAL_ROWS


def _phase_from_csv(raw: str):
    """Map a trial-master phase string ('PHASE3', 'PHASE1|PHASE2', ...) to a Phase.
    Highest phase token wins; unmappable -> None (orchestrator defaults to PH2)."""
    best, label = None, (str(raw).strip() or "N/A")
    for tok in re.split(r"[|/,]", str(raw).upper()):
        p = _PHASE_MAP.get(tok.strip().replace(" ", ""))
        if p and (best is None or p.value > best.value):
            best = p
    return best, label


def ctgov_pull(condition: str, term: str, *, max_pages: int = 5) -> list[dict]:
    """NON-DETERMINISTIC I/O, backed by the local trial-master CSV instead of the
    CT.gov API. Select rows whose conditions match `condition` (lung scope) and whose
    interventions match the drug `term` (which may carry ' OR ' alternatives), then
    emit the SAME dict shape ctgov_pull always returned. `max_pages` is ignored."""
    cond = (condition or "").lower()
    lung = any(k in cond for k in ("lung", "nsclc")) or not cond
    # term -> list of alternative word-sets; an alt matches if all its words are present
    alts = [[w for w in re.split(r"\W+", a.lower()) if w]
            for a in re.split(r"\bor\b", term or "", flags=re.I)]
    alts = [a for a in alts if a]
    out, seen = [], set()
    for r in _load_trial_master():
        conditions = (r.get("conditions") or "").lower()
        if lung and not any(k in conditions for k in ("lung", "nsclc", "non-small")):
            continue
        iv = (r.get("interventions") or "").lower()
        title = (r.get("trial") or "").lower()
        hay = iv + " " + title
        if alts and not any(all(w in hay for w in alt) for alt in alts):
            continue
        nct = (r.get("nct_number") or r.get("resolved_register_number")
               or r.get("register_number") or "").strip()
        if not nct or nct in seen:
            continue
        seen.add(nct)
        ph, ph_raw = _phase_from_csv(r.get("phase"))
        sd = (r.get("start_date") or r.get("study_first_posted") or "").strip() or None
        status = (r.get("overall_status") or "").strip().upper().replace(" ", "_") or None
        out.append({"nct": nct, "title": (r.get("trial") or "").strip(),
                    "phase": ph, "phase_raw": ph_raw,
                    "status": status,
                    "why_stopped": None,   # not carried in the trial-master CSV
                    "start_date": sd,
                    "start_year": int(sd[:4]) if sd and sd[:4].isdigit() else None,
                    "has_results": (r.get("has_results") or "").strip().lower()
                                   in ("yes", "true", "1")})
    return out


# --- 2b  [JUDGE]  outcome adjudication from the published readout -----------
# This is the agentic layer: an analyst/LLM reads the cited publication for a trial
# and returns its adjudicated endpoint outcome. Inlined here as the known NSCLC-2L
# readouts; swap in a live LLM-reading function with the SAME signature and nothing
# downstream changes — the deterministic scorer only consumes (endpoint_met, citation).
_OUTCOME_OVERLAY = {
    "NCT05215340": dict(endpoint="OS", met=False, combo=False, cite="TROPION-Lung01: OS not significant in broad 2L"),
    "NCT05089734": dict(endpoint="OS", met=False, combo=False, cite="EVOKE-01: OS miss vs docetaxel"),
    "NCT05555732": dict(endpoint="ORR", met=True, approved=True, combo=False, cite="dato-DXd EGFR-mutant NSCLC accelerated approval 2025 (TROPION-Lung05 pooled)"),
    "NCT04612751": dict(endpoint="ORR", met=True, combo=True, cite="TROPION-Lung04: dato-DXd + durvalumab, ORR signal (Ph1b)"),
    "NCT04644237": dict(endpoint="ORR", met=True, approved=True, combo=False, cite="DESTINY-Lung02: T-DXd ORR; FDA accelerated 2022"),
    "NCT03539536": dict(endpoint="ORR", met=True, approved=True, combo=False, cite="LUMINOSITY: teliso-V ORR 35%; FDA Emrelis 2025-05-14"),
    "NCT04294810": dict(endpoint="OS", met=False, combo=True, cite="SKYSCRAPER-01: final OS not met (AACR 2025)"),
    "NCT04154956": dict(endpoint="PFS", met=False, combo=False, cite="CARMEN-LC03: PFS co-primary missed; program ended 2023-12"),
    "NCT05983965": dict(endpoint="ORR", met=True, combo=False, cite="BL-B01D1 EGFRxHER3 ADC: early ORR signal NSCLC"),
}


def adjudicate_outcome(nct_id: str) -> Optional[dict]:
    """JUDGE plug-in. Return the cited, adjudicated outcome for a trial, or None if
    no published verdict exists (-> the trial stays non-decisive, never guessed)."""
    return _OUTCOME_OVERLAY.get(nct_id)


# --- 2c  [I/O]  patent landscape from the LOCAL claims-enriched IP-master CSV -
# No live PatSnap pull. The crowding signal comes entirely from the local FINALV2
# claims-enriched master. The registry still carries a TACD:(...) query string; we
# evaluate it as a boolean over each row's target + detected-modality columns, then
# keep only patents whose CLAIM text evidences the target (claims-grounded proxy).
_IP_ROWS: Optional[list[dict]] = None

# modality token (from the TACD query) -> the IP-master modality codes it covers.
# "antibody" is the broad antibody-format catch-all (mAb / bispecific / ADC / TCE).
_MODALITY_MAP = {
    "adc": {"ADC"}, "antibodydrugconjugate": {"ADC"},
    "bispecific": {"BISPECIFIC"},
    "tcellengager": {"BITE", "BISPECIFIC"}, "bite": {"BITE"},
    "antibody": {"MAB", "BISPECIFIC", "ADC", "BITE"},
    "mab": {"MAB"}, "monoclonal": {"MAB"},
}


def _load_ip_master() -> list[dict]:
    """Read (and cache) the IP-master CSV once. Preamble-tolerant: the exported file may
    be a multi-section workbook dump (README / Pull Metrics / Crowding Proxy / Enriched
    Master), so we locate the data table by its 'ip_master_id' header and read from there.
    A plain single-table CSV just starts at row 0. Missing file -> [] (flagged later)."""
    global _IP_ROWS
    if _IP_ROWS is None:
        try:
            with open(IP_MASTER_CSV, newline="", encoding="utf-8") as f:
                rows = list(csv.reader(f))                  # csv.reader handles multiline quoted fields
            start = next((i for i, r in enumerate(rows) if r and r[0].strip() == "ip_master_id"), 0)
            hdr = list(rows[start])
            while hdr and hdr[-1].strip() == "":            # trim trailing padding columns
                hdr.pop()
            ncol = len(hdr)
            _IP_ROWS = [dict(zip(hdr, (r + [""] * ncol)[:ncol]))
                        for r in rows[start + 1:] if any(c.strip() for c in r)]
            print(f"    [ip_csv] loaded {len(_IP_ROWS)} rows from "
                  f"{os.path.basename(IP_MASTER_CSV)} (data table @ row {start + 1})", file=sys.stderr)
        except Exception as e:
            print(f"    [ip_csv] WARN cannot read {IP_MASTER_CSV}: {e}", file=sys.stderr)
            _IP_ROWS = []
    return _IP_ROWS


def _norm_alnum(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def _tacd_groups(query: str) -> list[str]:
    """Pull the inner text of each top-level TACD:(...) clause, with balanced-paren
    handling (so a nested 'EGFR AND ("HER3" OR ERBB3)' stays intact)."""
    groups, i, key = [], 0, "TACD:("
    while True:
        j = query.find(key, i)
        if j < 0:
            break
        k, depth = j + len(key), 1
        start = k
        while k < len(query) and depth:
            if query[k] == "(":
                depth += 1
            elif query[k] == ")":
                depth -= 1
            k += 1
        groups.append(query[start:k - 1])
        i = k
    return groups


def _eval_group(group: str, atom_true) -> bool:
    """Evaluate one TACD clause as a boolean: terms joined by AND/OR (with parens),
    each atom resolved to True/False by `atom_true`. Atoms come from our own registry
    strings, so the tiny eval here is over a sanitised 0/1/and/or/paren expression."""
    expr = []
    for tok in re.findall(r'"[^"]+"|\(|\)|[A-Za-z0-9\-\.]+', group):
        up = tok.upper()
        if tok in "()":
            expr.append(tok)
        elif up == "AND":
            expr.append(" and ")
        elif up == "OR":
            expr.append(" or ")
        else:
            expr.append("1" if atom_true(tok.strip('"')) else "0")
    try:
        return bool(eval("".join(expr) or "0"))
    except Exception:
        return False


def _ip_row_matches(row: dict, query: str) -> bool:
    """True if a patent row satisfies the TACD query: clause-0 over target synonyms,
    remaining clauses over detected modalities (clauses joined by AND)."""
    groups = _tacd_groups(query)
    if not groups:
        return False
    ip_targets = [_norm_alnum(t) for t in str(row.get("target", "")).split("|") if _norm_alnum(t)]
    mods = {m.strip().upper() for m in (str(row.get("all_detected_modalities", "")) + "|"
            + str(row.get("modality_only", ""))).split("|") if m.strip()}

    def target_atom(a: str) -> bool:
        na = _norm_alnum(a)
        return any(na == it or (len(na) >= 3 and na in it) or (len(it) >= 3 and it in na)
                   for it in ip_targets)

    def modality_atom(a: str) -> bool:
        return bool(mods & _MODALITY_MAP.get(_norm_alnum(a), set()))

    # BISPEC-FIX: an AND-of-targets clause (e.g. 'EGFR AND HER3') must be satisfied by a
    # genuine pair-specific bispecific, NOT a multi-antigen 'list' patent that merely
    # enumerates both antigens among many. Exclude patents whose detected target set is
    # unfocused (> FOCUS_MAX_TARGETS distinct antigens) from multi-target (AND) cohorts.
    if re.search(r"\bAND\b", groups[0], re.I) and len(set(ip_targets)) > FOCUS_MAX_TARGETS:
        return False
    if not _eval_group(groups[0], target_atom):
        return False
    return all(_eval_group(g, modality_atom) for g in groups[1:])


def _ip_apd(row: dict) -> Optional[str]:
    """Application date as YYYYMMDD for cutoff/velocity windows; fall back to filing_year."""
    d = (row.get("application_date") or "").strip()
    if d.isdigit() and len(d) >= 4:
        return (d[:8]).ljust(8, "0")
    y = (row.get("filing_year") or "").strip()
    return (y + "0101") if y.isdigit() else None


# Claims-text crowding gate. In this enriched master the target/description columns are
# sparse and ~64% of rows have NO claims pulled (api_insufficient_balance). The only
# trustworthy evidence of real FTO crowding is the CLAIM text, so a patent counts toward
# density only if its claims actually evidence the target. `proxy_basis` is computed
# upstream per patent:
#   claim_target_and_modality / claim_target_only  -> target IS in the claims  (COUNTED)
#   claim_modality_only / description_only / not_detected                      (excluded)
_CLAIMS_GROUNDED_BASES = {"claim_target_and_modality", "claim_target_only"}


def _claims_grounded(row: dict) -> bool:
    """True only if the patent's CLAIM text evidences the target (claims were pulled and
    the target appears in the claims). This is the claims-text proxy for crowding."""
    if str(row.get("patsnap_claims_fulltext_status", "")).strip() != "pulled":
        return False
    basis = str(row.get("target_modality_crowding_proxy_basis", "")).strip()
    return basis in _CLAIMS_GROUNDED_BASES or \
        str(row.get("claim_mentions_target", "")).strip().lower() == "yes"


def _cohort_and_counts(base_query: str, *, cut: str):
    """Shared by both pulls: the claims-grounded cohort (stable APD,PN-desc order, deduped
    on publication_number), the count context, and the per-patent rows kept for claim text."""
    matched_broad = [r for r in _load_ip_master() if _ip_row_matches(r, base_query)]
    matched = [r for r in matched_broad if _claims_grounded(r)]

    def count(rows, lo, hi):
        return sum(1 for r in rows if (d := _ip_apd(r)) and lo <= d <= hi)

    total = count(matched, "19000101", cut)
    total_broad = count(matched_broad, "19000101", cut)
    coverage = round(total / total_broad, 3) if total_broad else None

    def _pn(r):
        return (r.get("publication_number") or "").strip()
    bounded = sorted((r for r in matched if (d := _ip_apd(r)) and d <= cut),
                     key=lambda r: (_ip_apd(r) or "", _pn(r)), reverse=True)
    cohort, seen, rows = [], set(), []
    for r in bounded:
        pid = _pn(r)
        if pid and pid not in seen:
            seen.add(pid); cohort.append(pid); rows.append(r)
    return matched, total, total_broad, coverage, cohort, rows, count


def _claim_text_for(row: dict) -> str:
    """Verbatim CLAIM excerpts for the epitope read (claims only — claims define scope)."""
    parts = []
    for c in ("claim_epitope_domain_competition_snippets", "claim_sequence_cdr_snippets"):
        s = (row.get(c) or "").strip()
        if s:
            parts.append(s)
    return "\n".join(parts)


def _norm_ws(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


# v3 RULE: the deterministic weighted FTO already produces X for every node, so the LLM is
# now only a one-way FLOOR CHECK. It is therefore OPTIONAL by default — with no client the
# run still completes (every node keeps its deterministic X; the floor check is skipped and
# flagged). Set REQUIRE_LLM = True to force the floor check to run (raises if no client).
REQUIRE_LLM = False
JUDGE_RETRIES = 3            # transient API failures are retried before giving up


def _make_llm_client():
    """anthropic client if ANTHROPIC_API_KEY is set, else None."""
    try:
        import anthropic
        if os.environ.get("ANTHROPIC_API_KEY"):
            return anthropic.Anthropic()
    except Exception as e:
        print(f"    [epitope] LLM client unavailable: {e}", file=sys.stderr)
    return None


_LLM_CLIENT = _make_llm_client()


def _resolve_client(client):
    """Return the client to use, enforcing the no-workaround rule."""
    cli = _LLM_CLIENT if client == "__use_default__" else client
    if cli is None and REQUIRE_LLM:
        raise RuntimeError(
            "v2 requires an LLM client to attempt the epitope read for every node "
            "(no workaround). Install `anthropic` and set ANTHROPIC_API_KEY, or pass a "
            "client explicitly. To intentionally bypass (NOT recommended), set "
            "locked_whitespace_workflow_csv_v2.REQUIRE_LLM = False.")
    return cli


def _judge_with_retry(target, recs, cli):
    """Attempt the judge up to JUDGE_RETRIES times. Returns (verdict, error_or_None).
    A persistent failure is reported as an error (-> UNRESOLVED_judge_error), never as a
    silently-clean abstain."""
    last = None
    for attempt in range(1, JUDGE_RETRIES + 1):
        try:
            return judge_epitope_crowding(target, recs, client=cli), None
        except Exception as e:                         # transient API/network error
            last = e
            print(f"    [epitope] {target}: judge attempt {attempt}/{JUDGE_RETRIES} failed: {e}",
                  file=sys.stderr)
    return None, last


def _faithful_evidence(verdict: EpitopeVerdict, claims_blob_norm: str) -> tuple:
    """Anti-hallucination guard: keep only cited quotes that ACTUALLY appear (normalised)
    in the supplied claim text. A quote the model invented is dropped here."""
    kept = []
    for e in verdict.evidence:
        q = _norm_ws(e.quote)
        probe = q[:60] if len(q) > 60 else q          # tolerate truncation of long quotes
        if probe and probe in claims_blob_norm:
            kept.append(e)
    return tuple(kept)


def _epitope_x(open_flag, conf: float):
    """Map the gated epitope verdict to the x-axis crowding (0 open .. 1 crowded).
    closed (binding face locked) -> high crowding; open (room remains) -> low; None -> UNRESOLVED.
    Confidence scales the distance from the 0.5 split."""
    if open_flag is None:
        return None, "UNRESOLVED"
    if open_flag:                                      # room remains -> LESS crowded
        return round(max(0.0, 0.5 - 0.4 * conf), 3), "open"
    return round(min(1.0, 0.5 + 0.4 * conf), 3), "closed"   # locked -> MORE crowded


def llm_crowding_pull(reg: dict, *, as_of: str, client="__use_default__") -> dict:
    """V2 x-axis. Runs the LLM epitope judge over the target's claims-grounded cohort
    (whitespace_pipeline.py contract), gates the verdict (verified_epitope_flag) AND
    checks every cited quote is really in the supplied claims, then maps open/closed to
    the x value. No claims / no client / unbacked verdict -> UNRESOLVED (never fabricated).
    Count context (cohort, coverage) is still reported, but x comes from the read.

    DATA SOURCE: reg["ip_query"] is a TACD string matched as a boolean over the LOCAL
    FINAL_NEWV2 master (IP_MASTER_CSV) — NOT a PatSnap API call. Claim text fed to the
    judge is the FINAL_NEWV2 claim-snippet columns. No network anywhere.

    NO-WORKAROUND RULE: the judge is attempted for EVERY node (with retries). A node is
    only UNRESOLVED after a real attempt, and is then flagged (epitope_resolved=False)."""
    base_query, target = reg["ip_query"], reg["target"]
    cut = as_of.replace("-", "")
    matched, total, total_broad, coverage, cohort, rows, _ = _cohort_and_counts(base_query, cut=cut)

    # assemble claim records (from FINAL_NEWV2 snippets) for the judge (stable order, capped)
    recs = []
    for r in rows:
        txt = _claim_text_for(r)
        if txt and len(recs) < JUDGE_MAX_PATENTS:
            recs.append({"patent_number": (r.get("publication_number") or "").strip(),
                         "claims_text": txt})

    if rows:
        top = rows[0]
        key = (top.get("publication_number") or "").strip() or None
        assignee = next((a.strip() for a in str(top.get("current_assignee", "")).split("|") if a.strip()), None)
    else:
        key = assignee = None

    # --- LLM REASONING (always attempted; no silent skip) + deterministic gates ---
    cli = _resolve_client(client)                  # raises if REQUIRE_LLM and no client
    verdict, err = _judge_with_retry(target, recs, cli)

    if err is not None:                            # attempt made, model failed after retries
        epi_open, epi_conf, x, status = None, None, None, "UNRESOLVED_judge_error"
        rationale = f"judge failed after {JUDGE_RETRIES} attempts: {err}"
        cited = ""
    else:
        blob = _norm_ws(" ".join(r["claims_text"] for r in recs))
        kept = _faithful_evidence(verdict, blob)
        if verdict.open is not None and not kept:
            verdict = EpitopeVerdict(open=None, confidence=0.0,
                                     rationale=(verdict.rationale or "") + " | GUARD: cited quotes not found in supplied claims -> abstain",
                                     evidence=())
        else:
            verdict = EpitopeVerdict(open=verdict.open, confidence=verdict.confidence,
                                     rationale=verdict.rationale, evidence=kept)
        epi_open, epi_conf = verified_epitope_flag(verdict)
        x, status = _epitope_x(epi_open, epi_conf)
        if x is None:
            status = "UNRESOLVED_no_claims" if not recs else "UNRESOLVED_abstain"
        rationale = (verdict.rationale or "")[:500]
        cited = ";".join(e.patent_number for e in verdict.evidence)

    resolved = x is not None
    if not resolved:
        print(f"    [epitope] UNRESOLVED {target}: {status} "
              f"(cohort={len(cohort)}, claim_recs={len(recs)})", file=sys.stderr)

    return {"patent_family_count": total, "patent_family_count_broad": total_broad,
            "claims_coverage": coverage,
            "cohort_patent_ids": cohort, "cohort_size": len(cohort),
            "key_patent": key, "blocking_assignee": assignee,
            "epitope_crowding": x, "crowding_status": status, "crowding_basis": "llm_epitope_judge",
            "epitope_resolved": resolved,
            "epitope_confidence": epi_conf if resolved else None,
            "epitope_rationale": rationale,
            "cited_claims": cited,
            # count proxy kept for reference / comparison with v1 (NOT the x-axis here)
            "target_modality_crowding": crowding_from_count(total, scale=CLAIMS_CROWDING_SCALE)}


# ===========================================================================
# v3 X-AXIS — DETERMINISTIC WEIGHTED FTO  +  one-way LLM floor check
# (the count-grounded layers are Part-1 deciders; the LLM only ratchets up)
# ===========================================================================
def _target_only_query(reg: dict) -> str:
    """The target clause of a node's TACD query, with the modality clause stripped — so it
    matches the target across ALL modalities (the L1 cross-modality epitope scope)."""
    groups = _tacd_groups(reg["ip_query"])
    return f"TACD:({groups[0]})" if groups else reg["ip_query"]


def _fto_count(pred, cut: str) -> int:
    """Count claims-grounded patents dated on/before `cut` that satisfy `pred`. Shared by
    L1 and L3 so every layer counts on the SAME grounded, cutoff-bounded universe."""
    n = 0
    for r in _load_ip_master():
        if not _claims_grounded(r):
            continue
        d = _ip_apd(r)
        if not d or d > cut:
            continue
        if pred(r):
            n += 1
    return n


def _is_locking_claim(r: dict) -> bool:
    """L1 signal: a broad epitope/domain/competition claim ('any antibody that binds X',
    'competes with reference mAb') locks the binding face for ALL modalities."""
    return str(r.get("claim_mentions_epitope_domain_or_competition", "")).strip().lower() == "yes"


def _in_indication(r: dict) -> bool:
    ind = (r.get("indications") or "").lower()
    return any(t in ind for t in _INDICATION_TERMS)


def _is_use_combo(r: dict) -> bool:
    cat = (r.get("category") or "").strip().lower()
    return any(c in cat for c in _USE_CATEGORIES)


def _has_io_context(r: dict) -> bool:
    blob = " ".join([str(r.get("indications", "")), str(r.get("category", "")),
                     str(r.get("title", "")),
                     str(r.get("target_modality_crowding_proxy_rationale", ""))]).lower()
    return any(t in blob for t in _IO_COMBO_TERMS)


def weighted_fto(reg: dict, *, cut: str):
    """DETERMINISTIC three-layer FTO. Returns (l1, l2, l3, x_det, n_l1, n_l2, n_l3).
    Same CSV + same query + same weights/scales -> identical output, every run.

      L1  target, ALL modalities, locking epitope claims      (cross-modality; HARD FLOOR)
      L2  target x THIS modality, claims-grounded             (format/conjugation density)
      L3  target x modality, in-indication use/combination    (+IO -> combination claim space)

      X_det = max(0.55*L1 + 0.25*L2 + 0.20*L3, L1)
    """
    tq = _target_only_query(reg)
    fq = reg["ip_query"]
    is_io = reg.get("configuration") == "+IO"

    n_l1 = _fto_count(lambda r: _ip_row_matches(r, tq) and _is_locking_claim(r), cut)
    n_l2 = _fto_count(lambda r: _ip_row_matches(r, fq), cut)

    def _p_l3(r):
        if not _ip_row_matches(r, fq) or not _in_indication(r) or not _is_use_combo(r):
            return False
        return _has_io_context(r) if is_io else True
    n_l3 = _fto_count(_p_l3, cut)

    l1 = fto_layer_density(n_l1, scale=FTO_SCALE_L1)
    l2 = fto_layer_density(n_l2, scale=FTO_SCALE_L2)
    l3 = fto_layer_density(n_l3, scale=FTO_SCALE_L3)
    blend = FTO_W_L1 * l1 + FTO_W_L2 * l2 + FTO_W_L3 * l3
    x_det = max(blend, l1) if FTO_L1_IS_HARD_FLOOR else blend
    x_det = round(min(1.0, max(0.0, x_det)), 3)
    return l1, l2, l3, x_det, n_l1, n_l2, n_l3


def _llm_floor_check(target: str, recs: list[dict], cli) -> tuple:
    """One-way LLM floor. Reads verbatim claim text; if it finds the binding face LOCKED
    (open=False, evidence-backed, gate-passed), it returns a crowding FLOOR. Open/abstain/
    no-client/error -> no floor (None). The LLM can never lower X — only raise it.
    Returns (floor_or_None, status, cited_patents, confidence_or_None)."""
    if cli is None:
        return None, "llm_floor_skipped_no_client", "", None
    if not recs:
        return None, "llm_floor_no_claims", "", None
    verdict, err = _judge_with_retry(target, recs, cli)
    if err is not None:
        return None, "llm_floor_error", "", None
    blob = _norm_ws(" ".join(r["claims_text"] for r in recs))
    kept = _faithful_evidence(verdict, blob)                 # anti-hallucination guard
    if verdict.open is not None and not kept:
        return None, "llm_floor_abstain_unbacked", "", None
    verdict = EpitopeVerdict(open=verdict.open, confidence=verdict.confidence,
                             rationale=verdict.rationale, evidence=kept)
    epi_open, epi_conf = verified_epitope_flag(verdict)      # deterministic provenance gate
    cited = ";".join(e.patent_number for e in verdict.evidence)
    if epi_open is False:                                    # locked -> raise the floor
        return round(min(1.0, 0.5 + 0.4 * epi_conf), 3), "llm_floor_locked", cited, epi_conf
    if epi_open is True:
        return None, "llm_floor_open_no_change", cited, epi_conf
    return None, "llm_floor_abstain", cited, epi_conf


def weighted_fto_pull(reg: dict, *, as_of: str, client="__use_default__") -> dict:
    """V3 x-axis producer. (1) DETERMINISTIC weighted three-layer FTO sets X_det (and the
    L1 hard floor). (2) The LLM acts ONLY as a one-way floor check: a gated, evidence-backed
    'locked' verdict can raise X to a minimum; everything else leaves X_det untouched.

        X_final = max(X_det, llm_floor_if_locked)

    No network for the deterministic part — it is pure over the local FINAL_NEWV2 master.
    With no LLM client the deterministic X still stands (floor check skipped + flagged)."""
    base_query, target = reg["ip_query"], reg["target"]
    cut = as_of.replace("-", "")

    # deterministic weighted FTO (the decider) ------------------------------------------
    l1, l2, l3, x_det, n_l1, n_l2, n_l3 = weighted_fto(reg, cut=cut)

    # cohort + count context (stable, deduped) — also feeds the floor check & the audit ---
    matched, total, total_broad, coverage, cohort, rows, _ = _cohort_and_counts(base_query, cut=cut)
    if rows:
        top = rows[0]
        key = (top.get("publication_number") or "").strip() or None
        assignee = next((a.strip() for a in str(top.get("current_assignee", "")).split("|") if a.strip()), None)
    else:
        key = assignee = None

    resolved = not (n_l1 == 0 and n_l2 == 0)
    if not resolved:
        x_det = None

    # one-way LLM floor check -------------------------------------------------------------
    cli = _resolve_client(client)                  # None unless a client is configured (REQUIRE_LLM=False)
    recs = []
    for r in rows:
        txt = _claim_text_for(r)
        if txt and len(recs) < JUDGE_MAX_PATENTS:
            recs.append({"patent_number": (r.get("publication_number") or "").strip(),
                         "claims_text": txt})
    floor, fstatus, cited, fconf = _llm_floor_check(target, recs, cli)

    x_final, floor_applied, floor_delta = x_det, False, 0.0
    if x_det is not None and floor is not None and floor > x_det:
        x_final, floor_applied, floor_delta = floor, True, round(floor - x_det, 3)

    if not resolved:
        status = "UNRESOLVED_no_patents"
    elif floor_applied:
        status = f"FLOORED_UP/{fstatus}"
    else:
        status = f"weighted_fto/{fstatus}"
    rationale = f"L1_lock={n_l1} L2_fmt={n_l2} L3_use={n_l3} | det={x_det} | {fstatus}"

    if not resolved:
        print(f"    [fto] UNRESOLVED {target}: no claims-grounded patents (L1={n_l1}, L2={n_l2})",
              file=sys.stderr)

    return {"epitope_crowding": x_final, "x_deterministic": x_det,
            "fto_l1_epitope": l1, "fto_l2_format": l2, "fto_l3_use": l3,
            "fto_l1_lock_count": n_l1, "fto_l3_use_count": n_l3,
            "floor_applied": floor_applied, "floor_delta": floor_delta,
            "crowding_status": status, "crowding_basis": "weighted_fto_v3(0.55/0.25/0.20,L1_floor)+llm_floor",
            "epitope_resolved": resolved,
            "epitope_confidence": fconf if floor_applied else None,
            "epitope_rationale": rationale, "cited_claims": cited,
            "patent_family_count": total, "patent_family_count_broad": total_broad,
            "claims_coverage": coverage,
            "cohort_patent_ids": cohort, "cohort_size": len(cohort),
            "key_patent": key, "blocking_assignee": assignee,
            # v1-style count proxy kept purely for reference / comparison (NOT the x-axis)
            "target_modality_crowding": crowding_from_count(total, scale=CLAIMS_CROWDING_SCALE)}


def crowding_pull(reg: dict, *, as_of: str, ps=None) -> dict:
    """DETERMINISTIC alternative (v1 behaviour): x = crowding_from_count(claims-grounded
    count). Kept for comparison; v2 uses llm_crowding_pull by default."""
    base_query = reg["ip_query"]
    cut = as_of.replace("-", "")
    matched, total, total_broad, coverage, cohort, rows, count = _cohort_and_counts(base_query, cut=cut)
    recent = count(matched, "20230101", cut)
    prior = count(matched, "20200101", "20221231")
    vel = round(recent / prior, 3) if (recent and prior) else None
    trend = None if vel is None else ("accelerating" if vel >= 1.0 else "steady" if vel >= 0.6 else "declining")
    if rows:
        top = rows[0]
        pn = (top.get("publication_number") or "").strip() or None
        assignee = next((a.strip() for a in str(top.get("current_assignee", "")).split("|") if a.strip()), None)
        yr = int((_ip_apd(top) or "0000")[:4]) or None
    else:
        pn = assignee = yr = None
    status = ("claims_text_empty" if not total
              else "claims_text_low_coverage" if (coverage is not None and coverage < 0.5)
              else "claims_text_density")
    x = crowding_from_count(total, scale=CLAIMS_CROWDING_SCALE)
    return {"patent_family_count": total, "patent_family_count_broad": total_broad,
            "claims_coverage": coverage, "patent_recent": recent, "patent_prior": prior,
            "patent_velocity_ratio": vel, "patent_velocity_trend": trend,
            "most_recent_filing_year": yr, "key_patent": pn, "blocking_assignee": assignee,
            "cohort_patent_ids": cohort, "cohort_size": len(cohort),
            "epitope_crowding": x,             # v2 uses epitope_crowding as the x field
            "target_modality_crowding": x,
            "crowding_status": status, "crowding_basis": "claims_text_count"}


# ===========================================================================
# PART 3 — ORCHESTRATOR   (fixed order; DI; only Part-1 functions decide)
# ===========================================================================
def run_node(reg: dict, *, as_of: str, as_of_year: int,
             pull_trials_fn: Callable, adjudicate_outcome_fn: Callable,
             pull_patents_fn: Callable) -> tuple[Node, list[dict]]:
    """One node, fixed sequence. Each external call is injected. The ONLY deciders are
    the deterministic Part-1 functions (date_filter, score_target, crowding, quadrant)."""
    # 1 [I/O]            pull the trial census
    raw = pull_trials_fn(TRIAL_CONDITION or TUMOR.lower(), reg["ctgov_term"])
    # 2 [DETERMINISTIC]  enforce the cutoff
    safe = date_filter(raw, as_of=as_of, date_key="start_date")
    # 3 [JUDGE]          adjudicate decisive outcomes from cited publications
    keep = set(reg.get("ncts", []))
    readouts, seen = [], set()
    for t in safe:
        ov = adjudicate_outcome_fn(t["nct"]) if t["nct"] in keep else None
        seen.add(t["nct"])
        readouts.append(_to_readout(reg, t, ov))
    for nct in keep - seen:                                  # curated readout missing from census
        ov = adjudicate_outcome_fn(nct)
        if ov:
            readouts.append(_to_readout(reg, {"nct": nct, "phase": Phase.PH3,
                            "status": "COMPLETED", "start_year": None, "has_results": False,
                            "why_stopped": None, "title": ""}, ov))
    # 4 [DETERMINISTIC]  clinical-performance score (y)
    tr = score_target(readouts, as_of_year=as_of_year, trial_volume=len(safe))
    # 5 [I/O]            patent landscape
    ip = pull_patents_fn(reg, as_of=as_of)
    # 6 + 7 [DETERMINISTIC] crowding transform (x) is inside crowding_pull; assemble + place
    n = Node(target=reg["target"], modality=reg["modality"], configuration=reg["configuration"],
             population=reg["population"], asset_examples=", ".join(reg["assets"]),
             biomarker=reg.get("biomarker"),
             validation_score=tr.performance, validation_label=tr.label,
             validation_confidence=tr.confidence,
             trials_total=len(raw), trials_kept_as_of=len(safe),
             trials_by_phase=_phase_dist(safe), ph3_failure=reg.get("ph3_fail"),
             discovery_complete=len(safe) > 0, ip_query=reg["ip_query"],
             citations={d["nct"]: d["citation"] for d in tr.detail if d["citation"]})
    for k, v in ip.items():
        setattr(n, k, v)
    assemble_node_score(n)
    return n, _trial_rows(reg, safe, adjudicate_outcome_fn, readouts)


def run_all(registry: list[dict], *, as_of=AS_OF, as_of_year=AS_OF_YEAR,
            pull_trials_fn=ctgov_pull, adjudicate_outcome_fn=adjudicate_outcome,
            pull_patents_fn=weighted_fto_pull) -> tuple[list[Node], Matrix2x2, list[dict]]:
    nodes, rows = [], []
    for reg in registry:
        n, tr = run_node(reg, as_of=as_of, as_of_year=as_of_year,
                         pull_trials_fn=pull_trials_fn,
                         adjudicate_outcome_fn=adjudicate_outcome_fn,
                         pull_patents_fn=pull_patents_fn)
        nodes.append(n); rows.extend(tr)
        print(f"  {n.label:42}  Y={str(n.validation_score):>5} [{n.validation_label}]  "
              f"X={str(n.epitope_crowding):>5} (det={n.x_deterministic}"
              f"{'+floor' if n.floor_applied else ''}) "
              f"L1/L2/L3={n.fto_l1_epitope}/{n.fto_l2_format}/{n.fto_l3_use} "
              f"[{n.crowding_status}]  trials={n.trials_kept_as_of}")
    return nodes, build_map1(nodes), rows


# --- small deterministic helpers used by the orchestrator ------------------
def _to_readout(reg, t, ov: Optional[dict]) -> Readout:
    return Readout(asset=reg["assets"][0], indication=f"{TUMOR} {LINE}",
                   phase=t["phase"] or Phase.PH2,
                   status="COMPLETED" if ov else (t.get("status") or "UNKNOWN"),
                   start_year=t.get("start_year"),
                   combo=(ov or {}).get("combo", reg["configuration"] == "+IO"),
                   primary_endpoint=(ov or {}).get("endpoint"),
                   endpoint_met=(ov or {}).get("met"),
                   approved_or_commercial=(ov or {}).get("approved", False),
                   why_stopped=(ov or {}).get("why", t.get("why_stopped")),
                   citation=(ov or {}).get("cite"),
                   nct_id=t["nct"], has_results=t.get("has_results", False),
                   superseded=(ov or {}).get("superseded", False))


def _phase_dist(trials) -> dict:
    d = {"PH1": 0, "PH2": 0, "PH3": 0, "APPROVAL": 0, "N/A": 0}
    for t in trials:
        d[t["phase"].name if t["phase"] else "N/A"] += 1
    return d


def _trial_rows(reg, safe, adjudicate_outcome_fn, readouts) -> list[dict]:
    keep = set(reg.get("ncts", []))
    rows = []
    for t in safe:
        decisive = t["nct"] in keep
        ov = adjudicate_outcome_fn(t["nct"]) if decisive else None
        ro = next((r for r in readouts if r.nct_id == t["nct"]), None)
        rows.append({"target": reg["target"], "modality": reg["modality"],
                     "configuration": reg["configuration"], "population": reg["population"],
                     "asset": reg["assets"][0], "nct_id": t["nct"], "title": t["title"],
                     "phase": t["phase"].name if t["phase"] else "N/A", "phase_raw": t.get("phase_raw"),
                     "overall_status": t.get("status"), "why_stopped": t.get("why_stopped"),
                     "has_results": t.get("has_results"), "start_year": t.get("start_year"),
                     "primary_endpoint": (ov or {}).get("endpoint", ""),
                     "endpoint_met": (ov or {}).get("met", ""),
                     "adjudicated_outcome": adjudicate(ro, AS_OF_YEAR).value if ro else "",
                     "decisive": decisive, "citation": (ov or {}).get("cite", "")})
    return rows


# ===========================================================================
# PART 4 — OUTPUTS   (graph + 3 spreadsheets)   [presentation only — no logic]
# ===========================================================================
def _write_table(path_xlsx, sheets):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        wb = Workbook(); wb.remove(wb.active)
        for name, (cols, rows) in sheets.items():
            ws = wb.create_sheet(name[:31]); ws.append(cols)
            for c in ws[1]:
                c.font = Font(bold=True)
            for r in rows:
                ws.append([r.get(c, "") for c in cols])
        wb.save(path_xlsx); print(f"  wrote {path_xlsx}")
    except ImportError:
        import csv
        for name, (cols, rows) in sheets.items():
            p = f"{path_xlsx[:-5]}__{name}.csv"
            with open(p, "w", newline="") as f:
                w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
                w.writeheader(); w.writerows(rows)
            print(f"  wrote {p}  (openpyxl absent -> CSV)")


def write_master(nodes):
    cols = ["target", "modality", "configuration", "population", "asset_examples",
            "node_score", "quadrant", "validation_score", "validation_label", "validation_confidence",
            "ph3_failure", "trials_total", "trials_kept_as_of",
            "biomarker", "biomarker_prevalence", "expression_prevalence_pct",
            "epitope_crowding", "x_deterministic", "floor_applied", "floor_delta",
            "fto_l1_epitope", "fto_l2_format", "fto_l3_use", "fto_l1_lock_count", "fto_l3_use_count",
            "epitope_resolved", "crowding_status", "crowding_basis", "epitope_confidence",
            "cited_claims", "epitope_rationale", "target_modality_crowding",
            "patent_family_count", "patent_family_count_broad", "claims_coverage", "cohort_size",
            "key_patent", "blocking_assignee", "ip_query"]
    _write_table(os.path.join(OUTDIR, "master_whitespace_nsclc_2l.xlsx"),
                 {"master": (cols, [asdict(n) for n in nodes])})


def write_trials(rows, nodes):
    cols = ["target", "modality", "configuration", "population", "asset", "nct_id", "title",
            "phase", "phase_raw", "overall_status", "why_stopped", "has_results", "start_year",
            "primary_endpoint", "endpoint_met", "adjudicated_outcome", "decisive", "citation"]
    scols = ["target", "modality", "configuration", "population", "trials_total", "trials_kept_as_of",
             "PH1", "PH2", "PH3", "APPROVAL", "N/A", "ph3_failure"]
    summ = [{"target": n.target, "modality": n.modality, "configuration": n.configuration,
             "population": n.population, "trials_total": n.trials_total,
             "trials_kept_as_of": n.trials_kept_as_of, **n.trials_by_phase,
             "ph3_failure": n.ph3_failure} for n in nodes]
    _write_table(os.path.join(OUTDIR, "clinical_trials_nsclc_2l.xlsx"),
                 {"trials": (cols, rows), "phase_summary": (scols, summ)})


def write_patent_recency(nodes):
    cols = ["target", "modality", "configuration", "population", "ip_query",
            "patent_family_count", "patent_family_count_broad", "claims_coverage",
            "patent_recent", "patent_prior", "patent_velocity_ratio", "patent_velocity_trend",
            "most_recent_filing_year", "target_modality_crowding", "crowding_basis", "cohort_size",
            "key_patent", "blocking_assignee", "crowding_status"]
    _write_table(os.path.join(OUTDIR, "patent_recency_nsclc_2l.xlsx"),
                 {"patent_recency": (cols, [asdict(n) for n in nodes])})


def write_patent_cohort(nodes):
    """One row per (node, patent_id) in stable rank order — the anchored cohort the
    claims-text pull iterates. Same IP-master + same query -> identical rows and order
    every run, so the cohort you read claims for never silently shifts."""
    cols = ["target", "modality", "configuration", "population", "rank", "patent_id"]
    rows = [{"target": n.target, "modality": n.modality, "configuration": n.configuration,
             "population": n.population, "rank": i, "patent_id": pid}
            for n in nodes for i, pid in enumerate(n.cohort_patent_ids, 1)]
    _write_table(os.path.join(OUTDIR, "patent_cohort_nsclc_2l.xlsx"),
                 {"patent_cohort": (cols, rows)})


# --- epitope-evidence extraction (anchored on each node's stable cohort) ------
# This is the INPUT layer for the (still-reserved) epitope read. It does NOT score
# anything: it pulls the CLAIM-level epitope/CDR language for every patent in a node's
# stable cohort, so the judge (e.g. epitope.py:judge_epitope_crowding) reads a fixed,
# rank-ordered set. Claims only — claims define legal scope; description is context.
_IP_BY_PN: Optional[dict] = None
_CLAIM_EVIDENCE_COLS = [
    "patsnap_claims_fulltext_status", "claims_text_chars",
    "claim_mentions_target", "claim_mentions_sequence_or_cdr",
    "claim_mentions_epitope_domain_or_competition",
    "claim_sequence_cdr_snippets", "claim_epitope_domain_competition_snippets",
]


def _ip_index_by_pn() -> dict:
    """publication_number -> enriched row (cached). Cohort ids index straight in."""
    global _IP_BY_PN
    if _IP_BY_PN is None:
        _IP_BY_PN = {}
        for r in _load_ip_master():
            pn = (r.get("publication_number") or "").strip()
            if pn and pn not in _IP_BY_PN:
                _IP_BY_PN[pn] = r
    return _IP_BY_PN


def _clean_cell(s, limit: int = 30000) -> str:
    """Strip control chars Excel rejects and truncate long claim snippets so the
    workbook always writes."""
    s = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", str(s or ""))
    return s if len(s) <= limit else s[:limit] + " …[truncated]"


def _yes(v) -> bool:
    return str(v or "").strip().lower() == "yes"


def write_epitope_evidence(nodes, m: "Matrix2x2"):
    """Two sheets: per-(node, cohort patent) claim evidence, and a per-node rollup of
    claim-level epitope/CDR hits. `shortlisted` marks the viable targets (top-2
    quadrants) so the epitope read can be run on those first."""
    idx = _ip_index_by_pn()
    sl = {n.label for n in m.shortlist()}

    cols = ["target", "modality", "configuration", "population", "shortlisted",
            "rank", "patent_id", "current_assignee", "application_date", "title"] + _CLAIM_EVIDENCE_COLS
    rows = []
    for n in nodes:
        for rank, pid in enumerate(n.cohort_patent_ids, 1):
            r = idx.get(pid, {})
            row = {"target": n.target, "modality": n.modality, "configuration": n.configuration,
                   "population": n.population, "shortlisted": n.label in sl, "rank": rank,
                   "patent_id": pid, "current_assignee": _clean_cell(r.get("current_assignee"), 500),
                   "application_date": r.get("application_date", ""),
                   "title": _clean_cell(r.get("title"), 500)}
            for c in _CLAIM_EVIDENCE_COLS:
                row[c] = _clean_cell(r.get(c))
            rows.append(row)

    scols = ["target", "modality", "configuration", "population", "shortlisted", "cohort_size",
             "claims_pulled", "claim_epitope_or_competition_hits", "claim_sequence_cdr_hits"]
    summ = []
    for n in nodes:
        crows = [idx.get(pid, {}) for pid in n.cohort_patent_ids]
        summ.append({"target": n.target, "modality": n.modality, "configuration": n.configuration,
                     "population": n.population, "shortlisted": n.label in sl,
                     "cohort_size": n.cohort_size,
                     "claims_pulled": sum(1 for r in crows if str(r.get("patsnap_claims_fulltext_status", "")).strip() == "pulled"),
                     "claim_epitope_or_competition_hits": sum(1 for r in crows if _yes(r.get("claim_mentions_epitope_domain_or_competition"))),
                     "claim_sequence_cdr_hits": sum(1 for r in crows if _yes(r.get("claim_mentions_sequence_or_cdr")))})

    _write_table(os.path.join(OUTDIR, "epitope_evidence_nsclc_2l.xlsx"),
                 {"epitope_evidence": (cols, rows), "epitope_signal": (scols, summ)})


def plot_map(m: Matrix2x2):
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        print("  [plot] matplotlib absent -> skipping PNG", file=sys.stderr); return
    col = {"TRUE WHITE SPACE": "#1a9850", "BATTLEGROUND": "#fdae61",
           "R&D TRAP": "#74add1", "RED FLAGS": "#d73027"}
    fig, ax = plt.subplots(figsize=(10, 8))
    ax.axhline(PERF_SPLIT, color="#999", lw=1); ax.axvline(CROWDING_SPLIT, color="#999", lw=1)
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.set_xlabel("patent / FTO crowding  (open -> crowded)")
    ax.set_ylabel("clinical performance  (failed -> validated)")
    ax.set_title(f"{TUMOR} {LINE} — Feasibility / Ownership 2x2  (as of {AS_OF})")
    for q, (hx, hy, ha) in {"TRUE WHITE SPACE": (.02, .97, "left"), "BATTLEGROUND": (.98, .97, "right"),
                            "R&D TRAP": (.02, .03, "left"), "RED FLAGS": (.98, .03, "right")}.items():
        ax.text(hx, hy, q, color=col[q], fontsize=9, ha=ha,
                va="top" if hy > .5 else "bottom", transform=ax.transAxes)
    for q in ALL_QUADRANTS:
        for n in m.cells[q]:
            ax.scatter(n.epitope_crowding, n.node_score, s=160, color=col[q],
                       marker="X" if n.ph3_failure else "o", edgecolors="black", linewidths=.8, zorder=3)
            ax.annotate(f"{n.target} {n.modality}\n({n.configuration}/{n.population})",
                        (n.epitope_crowding, n.node_score), textcoords="offset points",
                        xytext=(8, 6), fontsize=7.5)
    if m.unresolved:
        ax.text(0.5, -0.13, "UNRESOLVED (axis missing): " + ", ".join(n.label for n in m.unresolved),
                transform=ax.transAxes, ha="center", fontsize=7.5, color="#555")
    fig.tight_layout()
    out = os.path.join(OUTDIR, "map_nsclc_2l_2x2.png")
    fig.savefig(out, dpi=150, bbox_inches="tight"); print(f"  wrote {out}")


# ===========================================================================
# PART 5 — REGISTRY + main   (the locked NSCLC 2L set; TROP2 split 3 ways)
# ===========================================================================
REGISTRY = [
    dict(target="HER2", modality="ADC", configuration="mono", population="HER2-mutant",
         biomarker="HER2 activating mutation", assets=["trastuzumab deruxtecan"],
         ctgov_term="trastuzumab deruxtecan", ncts=["NCT04644237"],
         ip_query='TACD:("HER2" OR "ERBB2") AND TACD:(antibody OR "antibody-drug conjugate" OR ADC)'),
    dict(target="c-MET", modality="ADC", configuration="mono", population="c-Met-high",
         biomarker="c-Met high IHC / METex14", assets=["telisotuzumab vedotin", "capmatinib"],
         ctgov_term="telisotuzumab vedotin", ncts=["NCT03539536"],
         ip_query='TACD:("c-MET" OR "MET" OR "HGFR") AND TACD:(antibody OR ADC)'),
    dict(target="HER3", modality="ADC", configuration="mono", population="EGFR-mutant",
         biomarker="EGFR-mutant", assets=["patritumab deruxtecan"],
         ctgov_term="patritumab deruxtecan", ncts=[],
         ip_query='TACD:("HER3" OR "ERBB3") AND TACD:(antibody OR ADC)'),
    dict(target="EGFRxHER3", modality="bispecific ADC", configuration="mono", population="EGFR-mutant",
         biomarker="EGFR-mutant", assets=["BL-B01D1"],
         ctgov_term="BL-B01D1", ncts=["NCT05983965"],
         ip_query='TACD:(EGFR AND ("HER3" OR ERBB3)) AND TACD:(bispecific AND (antibody OR ADC))'),
    # TROP2 — one antigen, THREE nodes (the IO vs mono-ADC distinction the map turns on)
    dict(target="TROP2", modality="ADC", configuration="mono", population="driver+",
         biomarker="EGFR-mut / genomically-altered", assets=["datopotamab deruxtecan"],
         ctgov_term="datopotamab deruxtecan", ncts=["NCT05555732"],
         ip_query='TACD:("TROP2" OR "TACSTD2") AND TACD:(antibody OR ADC)'),
    dict(target="TROP2", modality="ADC", configuration="mono", population="all-comers",
         biomarker="none (unselected)", assets=["datopotamab deruxtecan", "sacituzumab govitecan"],
         ctgov_term="datopotamab deruxtecan OR sacituzumab govitecan", ncts=["NCT05215340", "NCT05089734"],
         ip_query='TACD:("TROP2" OR "TACSTD2") AND TACD:(antibody OR ADC)'),
    dict(target="TROP2", modality="ADC", configuration="+IO", population="combo",
         biomarker="PD-L1 (combo)", assets=["datopotamab deruxtecan + durvalumab"],
         ctgov_term="datopotamab deruxtecan durvalumab", ncts=["NCT04612751"],
         ip_query='TACD:("TROP2" OR "TACSTD2") AND TACD:(antibody OR ADC)'),
    dict(target="CEACAM5", modality="ADC", configuration="mono", population="CEACAM5-high",
         biomarker="CEACAM5-high", assets=["tusamitamab ravtansine"],
         ctgov_term="tusamitamab ravtansine", ncts=["NCT04154956"], ph3_fail=True,
         ip_query='TACD:("CEACAM5" OR "CEA") AND TACD:(antibody OR ADC)'),
    dict(target="TIGIT", modality="mAb", configuration="+IO", population="PD-L1",
         biomarker="PD-L1", assets=["tiragolumab"],
         ctgov_term="tiragolumab", ncts=["NCT04294810"], ph3_fail=True,
         ip_query='TACD:("TIGIT") AND TACD:(antibody)'),
    dict(target="DLL3", modality="TCE", configuration="mono", population="DLL3+",
         biomarker="DLL3 (~0-4% NSCLC vs ~70% SCLC)", assets=["tarlatamab"],
         ctgov_term="tarlatamab", ncts=[],
         ip_query='TACD:("DLL3") AND TACD:(antibody OR "T-cell engager" OR bispecific)'),
]


def _parse_args(argv=None):
    """Run-time scope overrides. Defaults reproduce the locked NSCLC 2L run exactly, so
    `python3 locked_whitespace_workflow_csv_v3_param.py` (no flags) == the original v3."""
    import argparse
    ap = argparse.ArgumentParser(
        description="Locked white-space workflow (v3, parameterized). Point the output "
                    "directory, tumor and cell/line at run time without editing the file.")
    ap.add_argument("--outdir", default=OUTDIR,
                    help="directory for ALL outputs (default: %(default)s)")
    ap.add_argument("--tumor", default=TUMOR,
                    help="tumor type, e.g. NSCLC, SCLC, mBC (default: %(default)s)")
    ap.add_argument("--line", "--cell", dest="line", default=LINE,
                    help="line / cell context, e.g. 1L, 2L, 3L (default: %(default)s)")
    ap.add_argument("--as-of", dest="as_of", default=AS_OF,
                    help="leakage cutoff YYYY-MM-DD (default: %(default)s)")
    ap.add_argument("--condition", default=TRIAL_CONDITION,
                    help="CT.gov condition scope passed to the trial pull; "
                         "default derives from --tumor (e.g. 'nsclc')")
    return ap.parse_args(argv)


if __name__ == "__main__":
    _args = _parse_args()
    # Apply run-time overrides to the module-level scope the workflow reads. These are
    # plain module globals here, so every Part-1/Part-3 function picks them up at call time.
    TUMOR, LINE = _args.tumor, _args.line
    AS_OF = _args.as_of
    AS_OF_YEAR = int(str(AS_OF)[:4]) if str(AS_OF)[:4].isdigit() else AS_OF_YEAR
    TRIAL_CONDITION = _args.condition
    OUTDIR = _args.outdir

    os.makedirs(OUTDIR, exist_ok=True)
    print(f"\nLOCKED white-space workflow — {TUMOR} {LINE}  (as of {AS_OF})\n" + "-" * 78)
    print(f"    [run] outdir={OUTDIR}  condition={TRIAL_CONDITION or TUMOR.lower()}", file=sys.stderr)

    # Wire the REAL adapters. Swap any of these three for a mock/LLM with the same
    # signature and the deterministic core is unaffected — that's the contract.
    # as_of / as_of_year are passed explicitly so the CLI overrides take effect (the
    # run_all defaults were bound at import time, before the overrides were parsed).
    nodes, m, trial_rows = run_all(
        REGISTRY,
        as_of=AS_OF, as_of_year=AS_OF_YEAR,
        pull_trials_fn=ctgov_pull,            # [I/O]   trial-master CSV (was ClinicalTrials.gov)
        adjudicate_outcome_fn=adjudicate_outcome,  # [JUDGE] published-readout adjudication
        pull_patents_fn=weighted_fto_pull,    # [x]   v3: deterministic weighted FTO + one-way LLM floor
    )

    print("\n" + "=" * 78 + f"\nMAP 1 — {TUMOR} {LINE} Feasibility / Ownership 2x2\n" + "=" * 78)
    for q in ALL_QUADRANTS:
        names = "  ".join(f"{n.target}/{n.configuration}/{n.population}"
                          f"{' [Ph3x]' if n.ph3_failure else ''}({n.node_score})" for n in m.cells[q])
        print(f"  {q:18}: {names or '-'}")
    print(f"  {'UNRESOLVED':18}: " + "  ".join(n.label for n in m.unresolved))
    print(f"\nsummary: {m.summary()}")
    print(f"shortlist (top-2 quadrants): {[n.label for n in m.shortlist()]}")

    print("\nCOVERAGE GAPS (flagged, not hidden):")
    for g in coverage_gaps(nodes):
        print(f"  - {g['node']}: {g['gaps']}")

    # NO-WORKAROUND FLAG: every node was attempted; any still-unresolved epitope is reported.
    unresolved = [n for n in nodes if not n.epitope_resolved]
    print("\n" + "!" * 78)
    if unresolved:
        print(f"UNRESOLVED EPITOPE — {len(unresolved)}/{len(nodes)} nodes attempted but NOT resolved:")
        for n in unresolved:
            print(f"  - {n.label}: {n.crowding_status}  (cohort={n.cohort_size})")
        print("These have NO x value and are excluded from the 2x2 placement until resolved.")
    else:
        print(f"EPITOPE RESOLVED — all {len(nodes)}/{len(nodes)} nodes returned a gated x value.")
    print("!" * 78)

    print("\nWriting outputs:")
    plot_map(m); write_master(nodes); write_trials(trial_rows, nodes); write_patent_recency(nodes)
    write_patent_cohort(nodes)
    write_epitope_evidence(nodes, m)
    print("\ndone.")
